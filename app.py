import os
import sqlite3
import hashlib
import shutil
import difflib
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, send_from_directory, jsonify, session, flash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
# Thêm vào app.py sau các import

import zipfile
import tempfile
import shutil

# ==================== BACKUP & RESTORE APIs ====================

@app.route('/api/backup/list')
def api_backup_list():
    """Lấy danh sách các bản sao lưu"""
    backups = []
    backup_dir = app.config.get('BACKUP_FOLDER', 'backups')
    
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir, exist_ok=True)
    
    for f in os.listdir(backup_dir):
        if f.startswith('tthc_backup_') and (f.endswith('.db') or f.endswith('.zip')):
            filepath = os.path.join(backup_dir, f)
            stat = os.stat(filepath)
            backups.append({
                'name': f,
                'size': stat.st_size,
                'date': datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            })
    
    backups.sort(key=lambda x: x['date'], reverse=True)
    return jsonify({'backups': backups})

@app.route('/api/backup/create', methods=['POST'])
def api_backup_create():
    """Tạo bản sao lưu mới"""
    try:
        backup_dir = app.config.get('BACKUP_FOLDER', 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = os.path.join(backup_dir, f"tthc_backup_{timestamp}.db")
        
        # Copy database
        if os.path.exists('tthc.db'):
            shutil.copy2('tthc.db', backup_file)
            
            # Nén file để tiết kiệm dung lượng
            zip_file = os.path.join(backup_dir, f"tthc_backup_{timestamp}.zip")
            with zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.write(backup_file, 'tthc.db')
            
            # Xóa file db gốc sau khi nén
            os.remove(backup_file)
            
            return jsonify({
                'success': True,
                'filename': f"tthc_backup_{timestamp}.zip",
                'message': 'Tạo bản sao lưu thành công'
            })
        else:
            return jsonify({'success': False, 'error': 'Không tìm thấy database'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/backup/restore', methods=['POST'])
def api_backup_restore():
    """Phục hồi dữ liệu từ bản sao lưu"""
    try:
        data = request.get_json()
        filename = data.get('filename')
        
        if not filename:
            return jsonify({'success': False, 'error': 'Thiếu tên file'})
        
        backup_dir = app.config.get('BACKUP_FOLDER', 'backups')
        filepath = os.path.join(backup_dir, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': 'File không tồn tại'})
        
        # Backup database hiện tại trước khi restore
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pre_restore_backup = os.path.join(backup_dir, f"pre_restore_backup_{timestamp}.db")
        if os.path.exists('tthc.db'):
            shutil.copy2('tthc.db', pre_restore_backup)
        
        # Giải nén hoặc copy file
        temp_db = os.path.join(tempfile.gettempdir(), 'restore_temp.db')
        
        if filename.endswith('.zip'):
            with zipfile.ZipFile(filepath, 'r') as zf:
                # Tìm file .db trong zip
                db_files = [f for f in zf.namelist() if f.endswith('.db')]
                if db_files:
                    zf.extract(db_files[0], tempfile.gettempdir())
                    extracted_path = os.path.join(tempfile.gettempdir(), db_files[0])
                    shutil.copy2(extracted_path, 'tthc.db')
                    os.remove(extracted_path)
                else:
                    return jsonify({'success': False, 'error': 'File zip không chứa database'})
        else:
            shutil.copy2(filepath, 'tthc.db')
        
        return jsonify({'success': True, 'message': 'Phục hồi dữ liệu thành công'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/backup/download/<path:filename>')
def api_backup_download(filename):
    """Tải xuống file sao lưu"""
    try:
        backup_dir = app.config.get('BACKUP_FOLDER', 'backups')
        safe_filename = os.path.basename(filename)
        filepath = os.path.join(backup_dir, safe_filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'File không tồn tại'}), 404
        
        return send_from_directory(backup_dir, safe_filename, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/backup/delete', methods=['POST'])
def api_backup_delete():
    """Xóa bản sao lưu"""
    try:
        data = request.get_json()
        filename = data.get('filename')
        
        if not filename:
            return jsonify({'success': False, 'error': 'Thiếu tên file'})
        
        backup_dir = app.config.get('BACKUP_FOLDER', 'backups')
        safe_filename = os.path.basename(filename)
        filepath = os.path.join(backup_dir, safe_filename)
        
        if os.path.exists(filepath):
            os.remove(filepath)
            return jsonify({'success': True, 'message': 'Đã xóa'})
        else:
            return jsonify({'success': False, 'error': 'File không tồn tại'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/backup/upload', methods=['POST'])
def api_backup_upload():
    """Tải lên file sao lưu và phục hồi"""
    try:
        if 'backup_file' not in request.files:
            return jsonify({'success': False, 'error': 'Không có file được chọn'})
        
        file = request.files['backup_file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Chưa chọn file'})
        
        filename = secure_filename(file.filename)
        backup_dir = app.config.get('BACKUP_FOLDER', 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        
        # Lưu file tạm thời
        temp_path = os.path.join(backup_dir, f"uploaded_{filename}")
        file.save(temp_path)
        
        # Backup hiện tại
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pre_restore_backup = os.path.join(backup_dir, f"pre_restore_backup_{timestamp}.db")
        if os.path.exists('tthc.db'):
            shutil.copy2('tthc.db', pre_restore_backup)
        
        # Phục hồi
        if filename.endswith('.zip'):
            with zipfile.ZipFile(temp_path, 'r') as zf:
                db_files = [f for f in zf.namelist() if f.endswith('.db')]
                if db_files:
                    zf.extract(db_files[0], tempfile.gettempdir())
                    extracted_path = os.path.join(tempfile.gettempdir(), db_files[0])
                    shutil.copy2(extracted_path, 'tthc.db')
                    os.remove(extracted_path)
                else:
                    os.remove(temp_path)
                    return jsonify({'success': False, 'error': 'File zip không chứa database'})
        elif filename.endswith('.db') or filename.endswith('.sqlite') or filename.endswith('.sqlite3'):
            shutil.copy2(temp_path, 'tthc.db')
        else:
            os.remove(temp_path)
            return jsonify({'success': False, 'error': 'Định dạng file không hỗ trợ'})
        
        # Xóa file tạm
        os.remove(temp_path)
        
        return jsonify({'success': True, 'message': 'Tải lên và phục hồi thành công'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
# Try to import PyPDF2 for PDF support (optional)
try:
    import PyPDF2
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'sotuphap_dongthap_secret_key_2026')

# Folders
UPLOAD_FOLDER = 'uploads'
QUYET_DINH_FOLDER = 'uploads/quyet_dinh'
BACKUP_FOLDER = 'backups'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['QUYET_DINH_FOLDER'] = QUYET_DINH_FOLDER
app.config['BACKUP_FOLDER'] = BACKUP_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(QUYET_DINH_FOLDER, exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)

# ==================== DECORATORS ====================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            flash('Bạn không có quyền truy cập!', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def check_quyen(quyen_name):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if session.get('role') == 'admin':
                return f(*args, **kwargs)
            if session.get(quyen_name):
                return f(*args, **kwargs)
            flash('Bạn không có quyền thực hiện chức năng này!', 'danger')
            return redirect(url_for('dashboard'))
        return decorated
    return decorator

# ==================== DATABASE ====================
def init_db():
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    
    # Users table
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        password TEXT,
        role TEXT DEFAULT 'user',
        fullname TEXT,
        email TEXT,
        created_at DATE
    )''')
    
    # Permissions table
    c.execute('''CREATE TABLE IF NOT EXISTS user_quyen (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        quyen_them_tthc INTEGER DEFAULT 0,
        quyen_sua_tthc INTEGER DEFAULT 0,
        quyen_xoa_tthc INTEGER DEFAULT 0,
        quyen_them_qd INTEGER DEFAULT 0,
        quyen_sua_qd INTEGER DEFAULT 0,
        quyen_xoa_qd INTEGER DEFAULT 0,
        quyen_bao_cao INTEGER DEFAULT 0,
        quyen_so_sanh INTEGER DEFAULT 0,
        FOREIGN KEY(user_id) REFERENCES users(id)
    )''')
    
    # User linh vuc
    c.execute('''CREATE TABLE IF NOT EXISTS user_linh_vuc (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        linh_vuc TEXT,
        FOREIGN KEY(user_id) REFERENCES users(id)
    )''')
    
    # TTHC table
    c.execute('''CREATE TABLE IF NOT EXISTS tthc (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ma_tthc TEXT UNIQUE,
        ten_tthc TEXT,
        linh_vuc TEXT,
        phi TEXT,
        le_phi TEXT,
        lien_thong_cung_cap TEXT,
        lien_thong_02_cap TEXT,
        phi_dia_gioi TEXT,
        dvc_toan_trinh TEXT,
        dvc_mot_phan TEXT,
        dvc_cung_cap_tt TEXT,
        dich_vu_bcci TEXT,
        ghi_chu TEXT,
        cap_thuc_hien TEXT,
        trang_thai_cong_khai TEXT,
        so_quyet_dinh TEXT,
        co_quan_thuc_hien TEXT,
        trang_thai TEXT DEFAULT 'da_cong_bo',
        thanh_phan_ho_so TEXT,
        thoi_gian_giai_quyet TEXT,
        can_cu_phap_ly TEXT,
        so_luong_da_xu_ly INTEGER DEFAULT 0,
        so_luong_dang_xu_ly INTEGER DEFAULT 0,
        created_at DATE
    )''')
    
    # Quyet dinh table
    c.execute('''CREATE TABLE IF NOT EXISTS quyet_dinh (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        so_quyet_dinh TEXT UNIQUE,
        ten_quyet_dinh TEXT,
        ngay_ban_hanh DATE,
        loai TEXT,
        mo_ta TEXT,
        file_dinh_kem TEXT,
        noi_dung_ocr TEXT,
        created_at DATE
    )''')
    
    # Create admin
    admin_password = hashlib.sha256('admin@123'.encode()).hexdigest()
    c.execute("SELECT COUNT(*) FROM users WHERE username = 'Admin'")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO users (username, password, role, fullname, email, created_at) VALUES (?,?,?,?,?,?)",
                  ('Admin', admin_password, 'admin', 'Quản trị viên', 'admin@dongthap.gov.vn', datetime.now().date()))
        c.execute("INSERT INTO user_quyen (user_id) VALUES (1)")
    
    # Import sample data if tthc empty
    c.execute("SELECT COUNT(*) FROM tthc")
    if c.fetchone()[0] == 0:
        sample_data = [
            ('TTHC001', 'Đăng ký khai sinh', 'Hộ tịch', 'Miễn phí', '0', '', '', '', 'X', '', '', '', '', 'xa', 'Đã công bố', '', 'UBND cấp xã', 'da_cong_bo', 'Giấy khai sinh, Giấy chứng sinh', '03 ngày', 'Luật Hộ tịch', 120, 15, datetime.now().date()),
            ('TTHC002', 'Cấp Căn cước công dân', 'Căn cước', '50.000', '0', '', '', '', 'X', '', '', '', '', 'tinh', 'Đã công bố', '', 'Công an tỉnh', 'da_cong_bo', 'Tờ khai CCCD, Ảnh thẻ', '07 ngày', 'Luật CCCD', 250, 30, datetime.now().date()),
            ('TTHC003', 'Đăng ký kinh doanh', 'Kinh doanh', '100.000', '0', '', '', '', 'X', '', '', '', '', 'tinh', 'Đã công bố', '', 'Sở KH&ĐT', 'da_cong_bo', 'Đơn đăng ký, Điều lệ', '03 ngày', 'Luật DN', 89, 12, datetime.now().date()),
        ]
        for item in sample_data:
            c.execute('''INSERT INTO tthc (ma_tthc, ten_tthc, linh_vuc, phi, le_phi, lien_thong_cung_cap,
                      lien_thong_02_cap, phi_dia_gioi, dvc_toan_trinh, dvc_mot_phan, dvc_cung_cap_tt,
                      dich_vu_bcci, ghi_chu, cap_thuc_hien, trang_thai_cong_khai, so_quyet_dinh,
                      co_quan_thuc_hien, trang_thai, thanh_phan_ho_so, thoi_gian_giai_quyet,
                      can_cu_phap_ly, so_luong_da_xu_ly, so_luong_dang_xu_ly, created_at)
                      VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', item)
    
    conn.commit()
    conn.close()

init_db()

# ==================== AUTH ====================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        hashed = hashlib.sha256(password.encode()).hexdigest()
        
        conn = sqlite3.connect('tthc.db')
        c = conn.cursor()
        c.execute("SELECT id, username, role, fullname FROM users WHERE username = ? AND password = ?", (username, hashed))
        user = c.fetchone()
        
        if user:
            session['user_id'] = user[0]
            session['username'] = user[1]
            session['role'] = user[2]
            session['fullname'] = user[3] or user[1]
            
            # Load permissions
            c.execute("SELECT * FROM user_quyen WHERE user_id = ?", (user[0],))
            quyen = c.fetchone()
            if quyen:
                session['quyen_them_tthc'] = quyen[2]
                session['quyen_sua_tthc'] = quyen[3]
                session['quyen_xoa_tthc'] = quyen[4]
                session['quyen_them_qd'] = quyen[5]
                session['quyen_sua_qd'] = quyen[6]
                session['quyen_xoa_qd'] = quyen[7]
                session['quyen_bao_cao'] = quyen[8]
                session['quyen_so_sanh'] = quyen[9]
            
            # Load linh vuc
            c.execute("SELECT linh_vuc FROM user_linh_vuc WHERE user_id = ?", (user[0],))
            session['linh_vuc'] = [row[0] for row in c.fetchall()]
            
            conn.close()
            flash(f'Chào mừng {session["fullname"]} trở lại!', 'success')
            return redirect(url_for('dashboard'))
        else:
            conn.close()
            return render_template('login.html', error='Sai tên đăng nhập hoặc mật khẩu!')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/doi_mat_khau', methods=['GET', 'POST'])
@login_required
def doi_mat_khau():
    if request.method == 'POST':
        old = request.form.get('old_password')
        new = request.form.get('new_password')
        confirm = request.form.get('confirm_password')
        
        if new != confirm:
            flash('Mật khẩu mới không khớp!', 'danger')
        elif len(new) < 6:
            flash('Mật khẩu phải có ít nhất 6 ký tự!', 'danger')
        else:
            conn = sqlite3.connect('tthc.db')
            c = conn.cursor()
            old_hash = hashlib.sha256(old.encode()).hexdigest()
            c.execute("SELECT password FROM users WHERE id = ?", (session['user_id'],))
            if c.fetchone()[0] != old_hash:
                flash('Mật khẩu cũ không đúng!', 'danger')
            else:
                new_hash = hashlib.sha256(new.encode()).hexdigest()
                c.execute("UPDATE users SET password = ? WHERE id = ?", (new_hash, session['user_id']))
                conn.commit()
                flash('Đổi mật khẩu thành công!', 'success')
                return redirect(url_for('dashboard'))
            conn.close()
    return render_template('doi_mat_khau.html')

# ==================== DASHBOARD ====================
@app.route('/')
@login_required
def dashboard():
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    
    linh_vuc_filter = ""
    params = []
    if session.get('role') != 'admin' and session.get('linh_vuc'):
        placeholders = ','.join(['?'] * len(session['linh_vuc']))
        linh_vuc_filter = f"WHERE linh_vuc IN ({placeholders})"
        params = session['linh_vuc']
    
    c.execute(f"SELECT COUNT(*) FROM tthc {linh_vuc_filter}", params)
    tong_tthc = c.fetchone()[0] or 0
    
    c.execute(f"SELECT SUM(so_luong_da_xu_ly) FROM tthc {linh_vuc_filter}", params)
    da_xu_ly = c.fetchone()[0] or 0
    
    c.execute(f"SELECT SUM(so_luong_dang_xu_ly) FROM tthc {linh_vuc_filter}", params)
    dang_xu_ly = c.fetchone()[0] or 0
    
    c.execute("SELECT COUNT(*) FROM quyet_dinh")
    so_qd = c.fetchone()[0] or 0
    
    # Linh vuc stats
    c.execute(f"SELECT linh_vuc, COUNT(*) FROM tthc WHERE linh_vuc IS NOT NULL AND linh_vuc != '' {linh_vuc_filter.replace('WHERE', 'AND') if linh_vuc_filter else ''} GROUP BY linh_vuc ORDER BY COUNT(*) DESC LIMIT 10", 
              params if linh_vuc_filter else [])
    linh_vuc_data = c.fetchall()
    linh_vuc_labels = [row[0] for row in linh_vuc_data]
    linh_vuc_values = [row[1] for row in linh_vuc_data]
    
    # DVC stats
    c.execute(f"SELECT COUNT(*) FROM tthc WHERE dvc_toan_trinh = 'X' {linh_vuc_filter.replace('WHERE', 'AND') if linh_vuc_filter else ''}", params)
    toan_trinh = c.fetchone()[0] or 0
    c.execute(f"SELECT COUNT(*) FROM tthc WHERE dvc_mot_phan = 'X' {linh_vuc_filter.replace('WHERE', 'AND') if linh_vuc_filter else ''}", params)
    mot_phan = c.fetchone()[0] or 0
    c.execute(f"SELECT COUNT(*) FROM tthc WHERE dvc_cung_cap_tt = 'X' {linh_vuc_filter.replace('WHERE', 'AND') if linh_vuc_filter else ''}", params)
    cung_cap_tt = c.fetchone()[0] or 0
    
    # Cap thuc hien stats
    c.execute(f"SELECT cap_thuc_hien, COUNT(*) FROM tthc WHERE cap_thuc_hien IS NOT NULL {linh_vuc_filter.replace('WHERE', 'AND') if linh_vuc_filter else ''} GROUP BY cap_thuc_hien", params)
    cap_data = c.fetchall()
    cap_map = {'tinh': 'Cấp tỉnh', 'xa': 'Cấp xã', 'bo': 'Cấp Bộ', 'dung_chung': 'Dùng chung', 'lien_thong': 'Liên thông'}
    cap_labels = [cap_map.get(row[0], row[0]) for row in cap_data]
    cap_values = [row[1] for row in cap_data]
    
    # Recent TTHC
    c.execute(f"SELECT id, ma_tthc, ten_tthc, linh_vuc FROM tthc {linh_vuc_filter} ORDER BY id DESC LIMIT 5", params)
    recent_tthc = c.fetchall()
    
    conn.close()
    
    stats = {'tong_tthc': tong_tthc, 'da_xu_ly': da_xu_ly, 'dang_xu_ly': dang_xu_ly, 'so_quyet_dinh': so_qd}
    
    return render_template('dashboard.html', stats=stats, linh_vuc_labels=linh_vuc_labels, linh_vuc_values=linh_vuc_values,
                          dvc_data=[toan_trinh, mot_phan, cung_cap_tt], cap_labels=cap_labels, cap_values=cap_values,
                          recent_tthc=recent_tthc)

# ==================== TTHC CRUD ====================
@app.route('/tthc')
@login_required
def tthc_list():
    page = request.args.get('page', 1, int)
    per_page = 10
    offset = (page - 1) * per_page
    
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    
    linh_vuc_filter = ""
    params = []
    if session.get('role') != 'admin' and session.get('linh_vuc'):
        placeholders = ','.join(['?'] * len(session['linh_vuc']))
        linh_vuc_filter = f"AND linh_vuc IN ({placeholders})"
        params = session['linh_vuc']
    
    c.execute(f"SELECT COUNT(*) FROM tthc WHERE trang_thai = 'da_cong_bo' {linh_vuc_filter}", params)
    total = c.fetchone()[0]
    
    c.execute(f"SELECT id, ma_tthc, ten_tthc, linh_vuc, cap_thuc_hien, co_quan_thuc_hien, trang_thai, so_luong_da_xu_ly, dvc_toan_trinh, dvc_mot_phan FROM tthc WHERE trang_thai = 'da_cong_bo' {linh_vuc_filter} ORDER BY id DESC LIMIT ? OFFSET ?", params + [per_page, offset])
    tthc_list = c.fetchall()
    conn.close()
    
    total_pages = (total + per_page - 1) // per_page
    return render_template('tthc_list.html', tthc_list=tthc_list, page=page, total_pages=total_pages)

@app.route('/tthc/them', methods=['GET', 'POST'])
@login_required
@check_quyen('quyen_them_tthc')
def tthc_create():
    if request.method == 'POST':
        conn = sqlite3.connect('tthc.db')
        c = conn.cursor()
        try:
            c.execute('''INSERT INTO tthc (ma_tthc, ten_tthc, linh_vuc, phi, le_phi, cap_thuc_hien, 
                      co_quan_thuc_hien, thanh_phan_ho_so, thoi_gian_giai_quyet, can_cu_phap_ly,
                      dvc_toan_trinh, dvc_mot_phan, dvc_cung_cap_tt, trang_thai, created_at)
                      VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                      (request.form.get('ma_tthc'), request.form.get('ten_tthc'), request.form.get('linh_vuc'),
                       request.form.get('phi'), request.form.get('le_phi'), request.form.get('cap_thuc_hien'),
                       request.form.get('co_quan_thuc_hien'), request.form.get('thanh_phan_ho_so'),
                       request.form.get('thoi_gian_giai_quyet'), request.form.get('can_cu_phap_ly'),
                       request.form.get('dvc_toan_trinh', ''), request.form.get('dvc_mot_phan', ''),
                       request.form.get('dvc_cung_cap_tt', ''), 'da_cong_bo', datetime.now().date()))
            conn.commit()
            flash('Thêm TTHC thành công!', 'success')
            return redirect(url_for('tthc_list'))
        except Exception as e:
            flash(f'Lỗi: {str(e)}', 'danger')
        finally:
            conn.close()
    
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("SELECT DISTINCT linh_vuc FROM tthc WHERE linh_vuc IS NOT NULL AND linh_vuc != ''")
    linh_vuc_list = [row[0] for row in c.fetchall()]
    conn.close()
    return render_template('tthc_form.html', tthc=None, linh_vuc_list=linh_vuc_list)

@app.route('/tthc/sua/<int:id>', methods=['GET', 'POST'])
@login_required
@check_quyen('quyen_sua_tthc')
def tthc_edit(id):
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    if request.method == 'POST':
        try:
            c.execute('''UPDATE tthc SET ma_tthc=?, ten_tthc=?, linh_vuc=?, phi=?, le_phi=?, 
                      cap_thuc_hien=?, co_quan_thuc_hien=?, thanh_phan_ho_so=?, thoi_gian_giai_quyet=?, 
                      can_cu_phap_ly=?, dvc_toan_trinh=?, dvc_mot_phan=?, dvc_cung_cap_tt=?
                      WHERE id=?''',
                      (request.form.get('ma_tthc'), request.form.get('ten_tthc'), request.form.get('linh_vuc'),
                       request.form.get('phi'), request.form.get('le_phi'), request.form.get('cap_thuc_hien'),
                       request.form.get('co_quan_thuc_hien'), request.form.get('thanh_phan_ho_so'),
                       request.form.get('thoi_gian_giai_quyet'), request.form.get('can_cu_phap_ly'),
                       request.form.get('dvc_toan_trinh', ''), request.form.get('dvc_mot_phan', ''),
                       request.form.get('dvc_cung_cap_tt', ''), id))
            conn.commit()
            flash('Cập nhật TTHC thành công!', 'success')
            return redirect(url_for('tthc_list'))
        except Exception as e:
            flash(f'Lỗi: {str(e)}', 'danger')
        finally:
            conn.close()
    c.execute("SELECT * FROM tthc WHERE id=?", (id,))
    tthc = c.fetchone()
    c.execute("SELECT DISTINCT linh_vuc FROM tthc WHERE linh_vuc IS NOT NULL AND linh_vuc != ''")
    linh_vuc_list = [row[0] for row in c.fetchall()]
    conn.close()
    return render_template('tthc_form.html', tthc=tthc, linh_vuc_list=linh_vuc_list)

@app.route('/tthc/xoa/<int:id>')
@login_required
@check_quyen('quyen_xoa_tthc')
def tthc_delete(id):
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("UPDATE tthc SET trang_thai = 'bai_bo' WHERE id=?", (id,))
    conn.commit()
    conn.close()
    flash('Đã xóa TTHC!', 'success')
    return redirect(url_for('tthc_list'))

@app.route('/tthc/chitiet/<int:id>')
@login_required
def tthc_detail(id):
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("SELECT * FROM tthc WHERE id=?", (id,))
    tthc = c.fetchone()
    conn.close()
    return render_template('tthc_detail.html', tthc=tthc)

# ==================== QUYET DINH ====================
@app.route('/quyet_dinh')
@login_required
def quyet_dinh_list():
    page = request.args.get('page', 1, int)
    per_page = 10
    offset = (page - 1) * per_page
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM quyet_dinh")
    total = c.fetchone()[0]
    c.execute("SELECT id, so_quyet_dinh, ten_quyet_dinh, ngay_ban_hanh, loai, file_dinh_kem FROM quyet_dinh ORDER BY id DESC LIMIT ? OFFSET ?", (per_page, offset))
    qd_list = c.fetchall()
    conn.close()
    total_pages = (total + per_page - 1) // per_page
    return render_template('quyet_dinh_list.html', quyet_dinh_list=qd_list, page=page, total_pages=total_pages)

@app.route('/quyet_dinh/them', methods=['GET', 'POST'])
@login_required
@check_quyen('quyen_them_qd')
def quyet_dinh_create():
    if request.method == 'POST':
        conn = sqlite3.connect('tthc.db')
        c = conn.cursor()
        try:
            file_dinh_kem = ''
            noi_dung_ocr = ''
            if 'file_dinh_kem' in request.files:
                file = request.files['file_dinh_kem']
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    filepath = os.path.join(app.config['QUYET_DINH_FOLDER'], filename)
                    file.save(filepath)
                    file_dinh_kem = filename
                    
                    # Simple PDF text extraction
                    if PDF_SUPPORT and filename.lower().endswith('.pdf'):
                        try:
                            with open(filepath, 'rb') as f:
                                reader = PyPDF2.PdfReader(f)
                                for page in reader.pages:
                                    text = page.extract_text()
                                    if text:
                                        noi_dung_ocr += text + '\n'
                        except:
                            noi_dung_ocr = ''
            
            c.execute('''INSERT INTO quyet_dinh (so_quyet_dinh, ten_quyet_dinh, ngay_ban_hanh, loai, mo_ta, file_dinh_kem, noi_dung_ocr, created_at)
                      VALUES (?,?,?,?,?,?,?,?)''',
                      (request.form.get('so_quyet_dinh'), request.form.get('ten_quyet_dinh'), request.form.get('ngay_ban_hanh'),
                       request.form.get('loai'), request.form.get('mo_ta'), file_dinh_kem, noi_dung_ocr, datetime.now().date()))
            conn.commit()
            flash('Thêm quyết định thành công!', 'success')
            return redirect(url_for('quyet_dinh_list'))
        except Exception as e:
            flash(f'Lỗi: {str(e)}', 'danger')
        finally:
            conn.close()
    return render_template('quyet_dinh_form.html', quyet_dinh=None)

@app.route('/quyet_dinh/sua/<int:id>', methods=['GET', 'POST'])
@login_required
@check_quyen('quyen_sua_qd')
def quyet_dinh_edit(id):
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    if request.method == 'POST':
        try:
            file_dinh_kem = request.form.get('existing_file', '')
            noi_dung_ocr = ''
            if 'file_dinh_kem' in request.files:
                file = request.files['file_dinh_kem']
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    filepath = os.path.join(app.config['QUYET_DINH_FOLDER'], filename)
                    file.save(filepath)
                    file_dinh_kem = filename
                    
                    if PDF_SUPPORT and filename.lower().endswith('.pdf'):
                        try:
                            with open(filepath, 'rb') as f:
                                reader = PyPDF2.PdfReader(f)
                                for page in reader.pages:
                                    text = page.extract_text()
                                    if text:
                                        noi_dung_ocr += text + '\n'
                        except:
                            noi_dung_ocr = ''
                    c.execute("UPDATE quyet_dinh SET noi_dung_ocr = ? WHERE id = ?", (noi_dung_ocr, id))
            
            c.execute('''UPDATE quyet_dinh SET so_quyet_dinh=?, ten_quyet_dinh=?, ngay_ban_hanh=?, 
                      loai=?, mo_ta=?, file_dinh_kem=? WHERE id=?''',
                      (request.form.get('so_quyet_dinh'), request.form.get('ten_quyet_dinh'),
                       request.form.get('ngay_ban_hanh'), request.form.get('loai'),
                       request.form.get('mo_ta'), file_dinh_kem, id))
            conn.commit()
            flash('Cập nhật quyết định thành công!', 'success')
            return redirect(url_for('quyet_dinh_list'))
        except Exception as e:
            flash(f'Lỗi: {str(e)}', 'danger')
        finally:
            conn.close()
    c.execute("SELECT * FROM quyet_dinh WHERE id=?", (id,))
    qd = c.fetchone()
    conn.close()
    return render_template('quyet_dinh_form.html', quyet_dinh=qd)

@app.route('/quyet_dinh/xoa/<int:id>')
@login_required
@check_quyen('quyen_xoa_qd')
def quyet_dinh_delete(id):
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("DELETE FROM quyet_dinh WHERE id=?", (id,))
    conn.commit()
    conn.close()
    flash('Đã xóa quyết định!', 'success')
    return redirect(url_for('quyet_dinh_list'))

@app.route('/quyet_dinh/chitiet/<int:id>')
@login_required
def quyet_dinh_detail(id):
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("SELECT * FROM quyet_dinh WHERE id=?", (id,))
    qd = c.fetchone()
    conn.close()
    return render_template('quyet_dinh_detail.html', qd=qd)

# ==================== SO SANH QUYET DINH ====================
@app.route('/so_sanh_quyet_dinh')
@login_required
def so_sanh_page():
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("SELECT id, so_quyet_dinh, ten_quyet_dinh FROM quyet_dinh ORDER BY ngay_ban_hanh DESC")
    qd_list = c.fetchall()
    conn.close()
    return render_template('so_sanh_quyet_dinh.html', quyet_dinh_list=qd_list)

@app.route('/api/so_sanh_quyet_dinh', methods=['POST'])
@login_required
def api_so_sanh():
    data = request.get_json()
    qd1_id = data.get('qd1_id')
    qd2_id = data.get('qd2_id')
    
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("SELECT * FROM quyet_dinh WHERE id = ?", (qd1_id,))
    qd1 = c.fetchone()
    c.execute("SELECT * FROM quyet_dinh WHERE id = ?", (qd2_id,))
    qd2 = c.fetchone()
    conn.close()
    
    if not qd1 or not qd2:
        return jsonify({'error': 'Không tìm thấy quyết định'}), 404
    
    text1 = qd1[7] if qd1[7] else qd1[5] if qd1[5] else ''
    text2 = qd2[7] if qd2[7] else qd2[5] if qd2[5] else ''
    
    diff = difflib.SequenceMatcher(None, text1, text2)
    diffs = []
    
    for tag, i1, i2, j1, j2 in diff.get_opcodes():
        if tag == 'replace':
            diffs.append({
                'type': 'replace',
                'position': {'qd1': {'start': i1, 'end': i2}, 'qd2': {'start': j1, 'end': j2}},
                'info': f'Nội dung khác nhau'
            })
        elif tag == 'insert':
            diffs.append({
                'type': 'insert',
                'position': {'qd2': {'start': j1, 'end': j2}},
                'info': f'Thêm mới trong QĐ2'
            })
        elif tag == 'delete':
            diffs.append({
                'type': 'delete',
                'position': {'qd1': {'start': i1, 'end': i2}},
                'info': f'Bị xóa trong QĐ2'
            })
    
    return jsonify({
        'qd1': {'so_quyet_dinh': qd1[1], 'ten_quyet_dinh': qd1[2], 'ngay_ban_hanh': qd1[3], 'loai': qd1[4], 'noi_dung': text1[:3000]},
        'qd2': {'so_quyet_dinh': qd2[1], 'ten_quyet_dinh': qd2[2], 'ngay_ban_hanh': qd2[3], 'loai': qd2[4], 'noi_dung': text2[:3000]},
        'diffs': diffs[:30]
    })

# ==================== BAO CAO ====================
@app.route('/bao_cao')
@login_required
def bao_cao():
    if session.get('role') != 'admin' and not session.get('quyen_bao_cao'):
        flash('Bạn không có quyền xem báo cáo!', 'danger')
        return redirect(url_for('dashboard'))
    
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    
    linh_vuc_filter = ""
    params = []
    if session.get('role') != 'admin' and session.get('linh_vuc'):
        placeholders = ','.join(['?'] * len(session['linh_vuc']))
        linh_vuc_filter = f"WHERE linh_vuc IN ({placeholders})"
        params = session['linh_vuc']
    
    c.execute(f"SELECT linh_vuc, COUNT(*) FROM tthc WHERE linh_vuc IS NOT NULL AND linh_vuc != '' {linh_vuc_filter.replace('WHERE', 'AND') if linh_vuc_filter else ''} GROUP BY linh_vuc ORDER BY COUNT(*) DESC", params)
    linh_vuc_stats = c.fetchall()
    
    c.execute(f"SELECT cap_thuc_hien, COUNT(*) FROM tthc WHERE cap_thuc_hien IS NOT NULL {linh_vuc_filter.replace('WHERE', 'AND') if linh_vuc_filter else ''} GROUP BY cap_thuc_hien", params)
    cap_stats = c.fetchall()
    
    c.execute(f"SELECT COUNT(*) FROM tthc WHERE dvc_toan_trinh = 'X' {linh_vuc_filter.replace('WHERE', 'AND') if linh_vuc_filter else ''}", params)
    toan_trinh = c.fetchone()[0] or 0
    c.execute(f"SELECT COUNT(*) FROM tthc WHERE dvc_mot_phan = 'X' {linh_vuc_filter.replace('WHERE', 'AND') if linh_vuc_filter else ''}", params)
    mot_phan = c.fetchone()[0] or 0
    
    conn.close()
    return render_template('bao_cao.html', linh_vuc_stats=linh_vuc_stats, cap_stats=cap_stats,
                          toan_trinh=toan_trinh, mot_phan=mot_phan)

# ==================== TIM KIEM ====================
@app.route('/tim_kiem', methods=['GET', 'POST'])
@login_required
def tim_kiem():
    results = []
    keyword = ''
    if request.method == 'POST':
        keyword = request.form.get('keyword', '').strip()
        if keyword:
            conn = sqlite3.connect('tthc.db')
            c = conn.cursor()
            
            linh_vuc_filter = ""
            params = [f'%{keyword}%', f'%{keyword}%', f'%{keyword}%']
            if session.get('role') != 'admin' and session.get('linh_vuc'):
                placeholders = ','.join(['?'] * len(session['linh_vuc']))
                linh_vuc_filter = f"AND linh_vuc IN ({placeholders})"
                params.extend(session['linh_vuc'])
            
            c.execute(f"SELECT id, ma_tthc, ten_tthc, linh_vuc, cap_thuc_hien, co_quan_thuc_hien FROM tthc WHERE (ten_tthc LIKE ? OR ma_tthc LIKE ? OR linh_vuc LIKE ?) {linh_vuc_filter} ORDER BY ten_tthc LIMIT 50", params)
            results = c.fetchall()
            conn.close()
    return render_template('tim_kiem.html', results=results, keyword=keyword)

# ==================== USER MANAGEMENT ====================
@app.route('/users')
@admin_required
def danh_sach_users():
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("SELECT id, username, role, fullname, email, created_at FROM users ORDER BY id")
    users = c.fetchall()
    conn.close()
    return render_template('users.html', users=users)

@app.route('/users/them', methods=['GET', 'POST'])
@admin_required
def them_user():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', 'user')
        fullname = request.form.get('fullname', '').strip()
        email = request.form.get('email', '').strip()
        
        hashed = hashlib.sha256(password.encode()).hexdigest()
        conn = sqlite3.connect('tthc.db')
        c = conn.cursor()
        try:
            c.execute("INSERT INTO users (username, password, role, fullname, email, created_at) VALUES (?,?,?,?,?,?)",
                      (username, hashed, role, fullname, email, datetime.now().date()))
            user_id = c.lastrowid
            c.execute("INSERT INTO user_quyen (user_id) VALUES (?)", (user_id,))
            conn.commit()
            flash('Thêm người dùng thành công!', 'success')
        except Exception as e:
            flash(f'Lỗi: {str(e)}', 'danger')
        finally:
            conn.close()
        return redirect(url_for('danh_sach_users'))
    return render_template('user_form.html', user=None)

@app.route('/users/sua/<int:id>', methods=['GET', 'POST'])
@admin_required
def sua_user(id):
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        role = request.form.get('role', 'user')
        fullname = request.form.get('fullname', '').strip()
        email = request.form.get('email', '').strip()
        new_password = request.form.get('new_password', '').strip()
        
        if new_password:
            hashed = hashlib.sha256(new_password.encode()).hexdigest()
            c.execute("UPDATE users SET username=?, password=?, role=?, fullname=?, email=? WHERE id=?",
                      (username, hashed, role, fullname, email, id))
        else:
            c.execute("UPDATE users SET username=?, role=?, fullname=?, email=? WHERE id=?",
                      (username, role, fullname, email, id))
        conn.commit()
        flash('Cập nhật người dùng thành công!', 'success')
        conn.close()
        return redirect(url_for('danh_sach_users'))
    
    c.execute("SELECT id, username, role, fullname, email FROM users WHERE id=?", (id,))
    user = c.fetchone()
    conn.close()
    return render_template('user_form.html', user=user)

@app.route('/users/xoa/<int:id>')
@admin_required
def xoa_user(id):
    if id == session.get('user_id'):
        flash('Không thể xóa chính mình!', 'danger')
        return redirect(url_for('danh_sach_users'))
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("DELETE FROM user_quyen WHERE user_id=?", (id,))
    c.execute("DELETE FROM user_linh_vuc WHERE user_id=?", (id,))
    c.execute("DELETE FROM users WHERE id=?", (id,))
    conn.commit()
    conn.close()
    flash('Đã xóa người dùng!', 'success')
    return redirect(url_for('danh_sach_users'))

# ==================== PHAN QUYEN ====================
@app.route('/phan_quyen')
@admin_required
def phan_quyen():
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("SELECT id, username, role, fullname FROM users WHERE role != 'admin' ORDER BY id")
    users = c.fetchall()
    c.execute("SELECT DISTINCT linh_vuc FROM tthc WHERE linh_vuc IS NOT NULL AND linh_vuc != ''")
    linh_vuc_list = [row[0] for row in c.fetchall()]
    if not linh_vuc_list:
        linh_vuc_list = ['HỘ TỊCH', 'LUẬT SƯ', 'CÔNG CHỨNG', 'ĐẤU GIÁ']
    conn.close()
    return render_template('phan_quyen.html', users=users, linh_vuc_list=linh_vuc_list)

@app.route('/api/get_permissions/<int:user_id>')
@admin_required
def get_permissions(user_id):
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    c.execute("SELECT * FROM user_quyen WHERE user_id = ?", (user_id,))
    q = c.fetchone()
    c.execute("SELECT linh_vuc FROM user_linh_vuc WHERE user_id = ?", (user_id,))
    linh_vuc = [row[0] for row in c.fetchall()]
    conn.close()
    
    return jsonify({
        'quyen_them_tthc': bool(q[2] if q else 0),
        'quyen_sua_tthc': bool(q[3] if q else 0),
        'quyen_xoa_tthc': bool(q[4] if q else 0),
        'quyen_them_qd': bool(q[5] if q else 0),
        'quyen_sua_qd': bool(q[6] if q else 0),
        'quyen_xoa_qd': bool(q[7] if q else 0),
        'quyen_bao_cao': bool(q[8] if q else 0),
        'quyen_so_sanh': bool(q[9] if q else 0),
        'linh_vuc': linh_vuc
    })

@app.route('/api/save_permissions', methods=['POST'])
@admin_required
def save_permissions():
    user_id = request.form.get('user_id')
    conn = sqlite3.connect('tthc.db')
    c = conn.cursor()
    
    c.execute('''UPDATE user_quyen SET quyen_them_tthc=?, quyen_sua_tthc=?, quyen_xoa_tthc=?,
                 quyen_them_qd=?, quyen_sua_qd=?, quyen_xoa_qd=?, quyen_bao_cao=?, quyen_so_sanh=?
                 WHERE user_id=?''',
              (1 if request.form.get('quyen_them_tthc') else 0,
               1 if request.form.get('quyen_sua_tthc') else 0,
               1 if request.form.get('quyen_xoa_tthc') else 0,
               1 if request.form.get('quyen_them_qd') else 0,
               1 if request.form.get('quyen_sua_qd') else 0,
               1 if request.form.get('quyen_xoa_qd') else 0,
               1 if request.form.get('quyen_bao_cao') else 0,
               1 if request.form.get('quyen_so_sanh') else 0,
               user_id))
    
    c.execute("DELETE FROM user_linh_vuc WHERE user_id = ?", (user_id,))
    for lv in request.form.getlist('linh_vuc'):
        if lv:
            c.execute("INSERT INTO user_linh_vuc (user_id, linh_vuc) VALUES (?,?)", (user_id, lv))
    
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ==================== BACKUP ====================
def backup_database():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(BACKUP_FOLDER, f"tthc_backup_{timestamp}.db")
    shutil.copy2('tthc.db', backup_file)
    return backup_file

def get_backup_list():
    backups = []
    for f in os.listdir(BACKUP_FOLDER):
        if f.startswith('tthc_backup_') and f.endswith('.db'):
            stat = os.stat(os.path.join(BACKUP_FOLDER, f))
            backups.append({'name': f, 'size': stat.st_size, 'date': datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")})
    return sorted(backups, key=lambda x: x['date'], reverse=True)

# Thay thế route backup cũ bằng route mới
@app.route('/backup')
def backup_page():
    return render_template('backup.html')

@app.route('/backup/create')
@admin_required
def create_backup():
    backup_database()
    flash('Đã tạo bản sao lưu!', 'success')
    return redirect(url_for('backup_list'))

@app.route('/backup/download/<filename>')
@admin_required
def download_backup(filename):
    return send_from_directory(BACKUP_FOLDER, filename, as_attachment=True)

@app.route('/backup/delete/<filename>')
@admin_required
def delete_backup(filename):
    os.remove(os.path.join(BACKUP_FOLDER, filename))
    flash('Đã xóa bản sao lưu!', 'success')
    return redirect(url_for('backup_list'))

# ==================== IMPORT EXCEL ====================
@app.route('/import-excel', methods=['GET', 'POST'])
@admin_required
def import_excel_page():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Không có file được chọn!', 'danger')
            return redirect(url_for('import_excel_page'))
        
        file = request.files['file']
        if file.filename == '':
            flash('Chưa chọn file!', 'danger')
            return redirect(url_for('import_excel_page'))
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            try:
                wb = load_workbook(filepath)
                ws = wb.active
                conn = sqlite3.connect('tthc.db')
                c = conn.cursor()
                success = 0
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1] and row[2]:  # Ma TTHC and Ten TTHC
                        try:
                            c.execute('''INSERT OR REPLACE INTO tthc 
                                      (ma_tthc, ten_tthc, linh_vuc, phi, le_phi, cap_thuc_hien, co_quan_thuc_hien, trang_thai, created_at)
                                      VALUES (?,?,?,?,?,?,?,?,?)''',
                                      (str(row[1]), str(row[2]), str(row[3]) if row[3] else '',
                                       str(row[4]) if row[4] else '', str(row[5]) if row[5] else '',
                                       str(row[15]) if len(row) > 15 and row[15] else '',
                                       str(row[16]) if len(row) > 16 and row[16] else '',
                                       'da_cong_bo', datetime.now().date()))
                            success += 1
                        except Exception as e:
                            print(f"Error: {e}")
                
                conn.commit()
                conn.close()
                os.remove(filepath)
                flash(f'Import thành công {success} dòng dữ liệu!', 'success')
            except Exception as e:
                flash(f'Lỗi import: {str(e)}', 'danger')
        else:
            flash('File không hợp lệ!', 'danger')
        return redirect(url_for('tthc_list'))
    
    return render_template('import_excel.html')

# ==================== STATIC FILES ====================
@app.route('/uploads/quyet_dinh/<filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['QUYET_DINH_FOLDER'], filename)

# ==================== RUN ====================
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
